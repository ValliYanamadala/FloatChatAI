import openpyxl

wb = openpyxl.load_workbook("argo_20_global_demo_extended.xlsx", data_only=True)
print("Sheet names:", wb.sheetnames)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n=================== SHEET: {sheetname} ===================")
    print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
    
    # Read first 15 rows
    rows = list(sheet.iter_rows(values_only=True))
    print(f"Total rows in memory: {len(rows)}")
    
    if sheetname == "README":
        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                print(f"Row {i+1}: {row}")
    else:
        # Headers
        headers = rows[0]
        print(f"\nHeaders ({len(headers)} columns):")
        for idx, h in enumerate(headers):
            print(f"  Col {idx+1}: {h}")
            
        print("\nFirst 3 data rows:")
        for r_idx in range(1, min(4, len(rows))):
            print(f"Row {r_idx+1}: {dict(zip(headers, rows[r_idx]))}")
            
        # Data types of non-null values per column
        print("\nColumn Data Types & Non-null counts:")
        for col_idx, h in enumerate(headers):
            col_values = [rows[r][col_idx] for r in range(1, len(rows))]
            non_null = [v for v in col_values if v is not None]
            types = set(type(v).__name__ for v in non_null)
            unique_count = len(set(non_null))
            sample_val = non_null[0] if non_null else None
            print(f"  {h} ({col_idx+1}): count={len(non_null)}/{len(col_values)}, unique={unique_count}, types={types}, sample={sample_val}")
